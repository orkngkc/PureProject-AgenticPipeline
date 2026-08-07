#!/usr/bin/env python3
"""Pipeline JSON çıktısını AGENT BAŞINA ayrı Excel dosyasına çevirir.

Kullanım:
  python3 export_excel.py output/pdf_test.json pdf
  python3 export_excel.py output/sade_test.json txt

Üretilen dosyalar (örn. prefix=pdf):
  output/excel/pdf_person_agent.xlsx     (figures)
  output/excel/pdf_place_agent.xlsx      (toponyms)
  output/excel/pdf_event_agent.xlsx      (events)
  ... her veri tipi için ayrı dosya.
"""
from __future__ import annotations

import json
import sys
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

# JSON anahtarı -> dosya adındaki agent adı
KEY_TO_AGENT = {
    "figures": "person",
    "toponyms": "place",
    "events": "event",
    "timeline": "timeline",
    "conquests": "conquest",
    "attributes": "attribute",
    "objects": "object",
    "relations": "relation",
}


def _cell(value):
    """Excel hücresi için değeri düzleştir (liste -> virgüllü metin)."""
    if isinstance(value, list):
        return ", ".join(str(x) for x in value)
    if value is None:
        return ""
    return value


def _columns(rows: list[dict]) -> list[str]:
    """Sütun sırası: ilk kaydın anahtarları + sonrakilerde çıkan ekstralar."""
    cols: list[str] = []
    for row in rows:
        for k in row:
            if k not in cols:
                cols.append(k)
    return cols


def export(json_path: str, prefix: str, out_dir: str = "output/excel") -> list[tuple[str, int]]:
    data = json.load(open(json_path, encoding="utf-8"))
    Path(out_dir).mkdir(parents=True, exist_ok=True)

    made: list[tuple[str, int]] = []
    for key, agent in KEY_TO_AGENT.items():
        rows = data.get(key, [])
        if not rows:                       # bu agent bu dosyada boşsa atla
            continue

        wb = Workbook()
        ws = wb.active
        ws.title = agent

        cols = _columns(rows)
        ws.append(cols)                    # başlık satırı
        for c in range(1, len(cols) + 1):  # başlığı kalın yap
            ws.cell(row=1, column=c).font = Font(bold=True)

        for row in rows:
            ws.append([_cell(row.get(c, "")) for c in cols])

        # Sütun genişliklerini içeriğe göre kabaca ayarla
        for i, col in enumerate(cols, start=1):
            width = max(len(str(col)),
                        *(len(str(_cell(r.get(col, "")))) for r in rows))
            ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = min(width + 2, 60)

        fname = f"{prefix}_{agent}_agent.xlsx"
        wb.save(Path(out_dir) / fname)
        made.append((fname, len(rows)))

    return made


if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Kullanım: python3 export_excel.py <json_yolu> <prefix>")
        sys.exit(1)
    files = export(sys.argv[1], sys.argv[2])
    print(f"{len(files)} Excel dosyası üretildi (output/excel/):")
    for name, n in files:
        print(f"   {name:32} ({n} satır)")
